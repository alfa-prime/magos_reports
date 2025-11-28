import io
import json

from openpyxl import Workbook
from async_lru import alru_cache
from openpyxl.styles import PatternFill

from app.core.decorators import log_and_catch
from app.core import ORGS_MAPPER, PAY_TYPE_MAPPER, logger
from app.service.gateway.gateway import GatewayService
from app.service.tool.tool import auto_cells_width, set_column_width, align_row_center, align_column_center


@alru_cache(maxsize=1000)
async def _fetch_job_data(person_id: str, gateway_service: GatewayService):
    payload = {
        "params": {
            "c": "Common",
            "m": "loadPersonData"
        },
        "data": {
            "onExpand": "true",
            "Person_id": person_id,
            "LoadShort": "false",
            "mode": "PersonInfoPanel",
        }
    }
    logger.info(f"Получение сведений о работе для person_id: '{person_id}'")

    response_json = await gateway_service.make_request(method="post", json=payload)
    job_id = response_json[0].get("JobOrg_id", "")
    job_name = response_json[0].get("Person_Job", "")
    soc_status = response_json[0].get("SocStatus_Name", "")

    if soc_status == "Работает" and not job_name:
        job_name = "В ЕВМИАС отсутствуют данные о месте работы"

    return {"job_id": job_id, "job_name": job_name, "soc_status": soc_status}


@alru_cache(maxsize=1000)
async def _fetch_usluga_code(usluga_complex_code_name: str, gateway_service: GatewayService) -> dict:
    payload = {
        "params": {
            "c": "UslugaComplex",
            "m": "loadUslugaContentsGrid"
        },
        "data": {
            "UslugaComplex_CodeName": usluga_complex_code_name,
            "object": "UslugaComplex",
            "UslugaComplex_pid": "3010101000029801",
            "contents": 2
        }
    }

    logger.info(f"Получение кода для услуги '{usluga_complex_code_name}'")
    response_json = await gateway_service.make_request(method="post", json=payload)
    return response_json[0].get("UslugaComplex_Code")


@alru_cache(maxsize=1000)
async def _fetch_pay_type(evn_direction_id: str, gateway_service: GatewayService):
    payload = {
        "params": {
            "c": "EvnLabRequest",
            "m": "load"
        },
        "data": {
            "EvnDirection_id": evn_direction_id,
            "delDocsView": 0,
        }
    }

    logger.info(f"Получение типа оплаты услуги для evn_direction_id: '{evn_direction_id}'")

    response_json = await gateway_service.make_request(method="post", json=payload)
    pay_type_id = response_json[0].get("PayType_id", "")
    return PAY_TYPE_MAPPER.get(pay_type_id, "")


async def _fetch_source_data(start_date: str, end_date: str, gateway_service: GatewayService):
    payload = {
        "params": {
            "c": "EvnLabRequest",
            "m": "loadEvnLabRequestList"
        },
        "data": {
            "EvnStatus_id": 2,
            "MedServiceType_SysNick": "reglab",
            "MedService_id": "3010101000015552",
            "fit": 1,
            "begDate": start_date,
            "endDate": end_date,
            "filterWorkELRByDate": 1,
            "filterDoneELRByDate": 1,
            "formMode": "false",
        }
    }
    response = await gateway_service.make_request(method="post", json=payload)
    logger.info("Исходные данные успешно получены")
    return response.get("data", "")  # noqa


@log_and_catch()
async def process_invitro_list(start_date: str, end_date: str, gateway_service: GatewayService):
    source_data = await _fetch_source_data(start_date, end_date, gateway_service)

    book = Workbook()
    sheet = book.active
    invalid_fill = PatternFill(start_color="FFE4E1", end_color="FFE4E1", fill_type="solid")

    titles = ["Фамилия", "Имя", "Отчество", "ДР", "Соц.статус", "Таможня/УФССП", "Место работы",
              "Дата услуги", "Вид оплаты", "Код услуги", "Услуга"]
    sheet.append(titles)

    for item in source_data:
        person_id = item.get("Person_id", "")

        services_list = json.loads(item.get("EvnLabRequest_UslugaName", ""))

        evn_direction_id = item.get("EvnDirection_id")
        pay_type = await _fetch_pay_type(evn_direction_id, gateway_service)

        surname = item.get("Person_Surname", "").title()
        first_name = item.get("Person_Firname", "").title()
        middle_name = item.get("Person_Secname", "").title()
        birthday = item.get("Person_Birthday", "")
        service_date = item.get("TimetableMedService_Date", "")

        for service in services_list:
            service_name = service.get("UslugaComplex_Name")
            service_code = await _fetch_usluga_code(service_name, gateway_service)

            job_data = await _fetch_job_data(person_id, gateway_service)
            job_id = job_data.get("job_id", "")
            job_name = job_data.get("job_name", "")
            soc_status = job_data.get("soc_status", "")

            job_category = ""
            if job_id in ORGS_MAPPER:
                job_category = ORGS_MAPPER[job_id]

            row_data = [
                surname,
                first_name,
                middle_name,
                birthday,
                soc_status,
                job_category,
                job_name,
                service_date,
                pay_type,
                service_code,
                service_name,
            ]

            sheet.append(row_data)

            # помечаем строки где нет данных о месте работы
            if job_name == "В ЕВМИАС отсутствуют данные о месте работы":
                current_row = sheet.max_row
                for cell in sheet[current_row]:
                    cell.fill = invalid_fill

    # автоширина всех колонок
    auto_cells_width(sheet)

    # задаем ширину колонки принудительно
    set_column_width(sheet=sheet, column_letter="G", width=45)

    # выравнивание первой строки по центру
    rows_to_align = [1]
    align_row_center(sheet, rows_to_align)

    # выравнивание колонок по центру
    columns_to_align = ["D", "H"]
    align_column_center(sheet, columns_to_align)

    # включить АВТОФИЛЬТР
    # sheet.dimensions вернет строку вида "A1:K150"
    sheet.auto_filter.ref = sheet.dimensions

    output_stream = io.BytesIO()
    book.save(output_stream)
    output_stream.seek(0)

    return output_stream
