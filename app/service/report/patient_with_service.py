from __future__ import annotations
import io
import asyncio
import json
from typing import List
from fastapi import HTTPException
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from async_lru import alru_cache
from tenacity import retry, stop_after_attempt, wait_fixed, retry_if_exception
from openpyxl import Workbook
from pydantic import BaseModel
from datetime import date, datetime
import warnings

from app.core import logger, PAY_TYPE_MAPPER
from app.service.gateway.gateway import GatewayService
from app.model.patient_with_services import PatientServiceRow
from app.service.tool.tool import auto_cells_width, set_column_width, align_row_center, align_column_center


@alru_cache(maxsize=1000)
async def _cached_search_hosp(gateway_service: GatewayService, card_number: str, hosp_start_date: str) -> dict:
    payload = {
        "params": {"c": "Search", "m": "searchData"},
        "data": {
            "PersonPeriodicType_id": 1,
            "SearchFormType": "EvnPS",
            "EvnPS_NumCard": card_number,
            "EvnSection_setDate_Range": f"{hosp_start_date} - {hosp_start_date}",
            "Date_Type": 1,
            "SearchType_id": 1,
            "PersonCardStateType_id": 1,
        }
    }
    return await gateway_service.make_request(method="post", json=payload)


@alru_cache(maxsize=1000)
async def _cached_search_hosp_services(gateway_service: GatewayService, hosp_id: str) -> list[dict]:
    payload = {
        "params": {"c": "EvnUsluga", "m": "loadEvnUslugaGrid"},
        "data": {
            "pid": hosp_id,
            "parent": "EvnPS",
        }
    }
    return await gateway_service.make_request(method="post", json=payload)


def _unmerge_and_fill_sheet(sheet: Worksheet):
    """
    Вспомогательная функция: Снимает объединение ячеек и заполняет их
    значением из верхней левой ячейки бывшего диапазона.
    """
    # Копируем список, так как будем менять структуру листа
    merged_ranges = list(sheet.merged_cells)

    for merged_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_range.bounds
        # Берем значение из главной ячейки
        top_left_value = sheet.cell(row=min_row, column=min_col).value
        # Снимаем объединение
        sheet.unmerge_cells(str(merged_range))
        # Заполняем все ячейки диапазона этим значением
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                sheet.cell(row=row, column=col).value = top_left_value


def _process_excel_sync(content: bytes) -> List[PatientServiceRow]:
    if not content.startswith(b'\x50\x4b\x03\x04'):
        # Если это не ZIP, попробуем понять, что это
        preview = content[:200].decode('utf-8', errors='ignore')
        raise ValueError(f"Полученный файл не является XLSX (ZIP). Начало файла: {preview}")

    input_stream = io.BytesIO(content)

    with warnings.catch_warnings():
        warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")
        workbook = load_workbook(input_stream, data_only=True)

    # 1. Снимаем объединения на всех листах
    for sheet in workbook.worksheets:
        _unmerge_and_fill_sheet(sheet)

    # 2. Склеиваем листы в один (если их > 1)
    main_sheet = workbook.worksheets[0]
    if len(workbook.worksheets) > 1:
        for sheet in workbook.worksheets[1:]:
            for row in sheet.iter_rows(min_row=4, values_only=True):
                if any(row):
                    main_sheet.append(row)
            workbook.remove(sheet)

    # Парсинг данных и удаление дублей
    result_data = []
    seen_rows = set()

    # Начинаем с 6-й строки
    for row in main_sheet.iter_rows(min_row=6, values_only=True):
        # Пропуски
        if not any(row): continue
        if row[0] and "итого" in str(row[0]).lower(): break

        # Удаление дублей (проверяем сырой кортеж)
        if row in seen_rows:
            continue
        seen_rows.add(row)

        try:
            patient_record = PatientServiceRow.from_row(row)
            result_data.append(patient_record)

        except Exception as e:
            logger.warning(f"Ошибка парсинга строки Excel: {e}")
            continue

    return result_data


async def get_list_patients_with_services(
        start_date: str,
        end_date: str,
        gateway_service: GatewayService
) -> List[PatientServiceRow]:
    report_params = (
        f"paramLpu=13102423&"
        f"__isnull=paramLpuBuilding&"
        f"__isnull=paramLpuUnit&"
        f"__isnull=paramLpuSection&"
        f"__isnull=paramMedStaffFact&"
        f"paramBegDate={start_date}&"
        f"paramEndDate={end_date}&"
        f"__isnull=paramUslugaClass&"
        f"paramDataOtch=1&"
        f"paramOnSection=1&"
        f"__isnull=paramPrehospRefuse&"
        f"param_RegionCode=301&"
        f"param_pmuser_id=461432278617"
    )

    payload = {
        "path": "/",
        "method": "GET",
        "params": {
            "c": "ReportRun",
            "m": "Run",
            "Report_id": "32430",
            "Report_Params": report_params,
            "__format": "xlsx",
            "__asattachment": "true"
        },
    }

    url = "/gateway/download"

    try:
        file_bytes = await gateway_service.download(
            url=url,
            method="POST",
            json=payload
        )

        logger.info(f"[CLIENT] Получено {len(file_bytes)} байт.")

    except Exception as e:
        logger.error(f"[CLIENT] Ошибка сети: {e}")
        raise HTTPException(503, "Не удалось связаться со шлюзом")

    try:
        # Запускаем в потоке
        result = await asyncio.to_thread(_process_excel_sync, file_bytes)

        for record in result:
            card_number = record.card_number
            hosp_start_date = record.start_date.strftime("%d.%m.%Y")
            service_date = record.service_date.strftime("%d.%m.%Y")
            try:
                patient_hosp_response = await _cached_search_hosp(gateway_service, card_number, hosp_start_date)
                hosp_data = patient_hosp_response.get("data")

                # Проверка: нашли ли госпитализацию?
                if not hosp_data or not isinstance(hosp_data, list):
                    logger.warning(f"Госпитализация не найдена для карты {card_number}")
                    continue

                patient_hosp_id = hosp_data[0].get("EvnPS_id")

                services_response = await _cached_search_hosp_services(gateway_service, patient_hosp_id)

                # Проверка: есть ли услуги?
                if not services_response or not isinstance(services_response, list):
                    continue

                if record.full_name == 'АБРОСИМОВА ЛЮДМИЛА ВИКТОРОВНА':
                    with open('./logs/temp.json', 'w', encoding='utf8') as file:
                        json.dump(services_response, file, indent=2, ensure_ascii=False)

                for each in services_response:
                    api_date = each.get("EvnUsluga_setDate", "")
                    api_code = each.get("Usluga_Code", "")

                    if api_code == record.service_code and api_date == service_date:
                        pay_type_id = each.get("PayType_id", "")
                        pay_source_name = PAY_TYPE_MAPPER.get(str(pay_type_id))

                        if pay_source_name:
                            record.service_payment_source = pay_source_name
                        else:
                            record.service_payment_source = f"Неизвестный id типа оплаты ({pay_type_id})"
                        break

            except Exception as e:
                logger.error(f"Ошибка при обогащении данных для карты {card_number}: {e}")
                continue

        return result

    except Exception as e:
        logger.error(f"[CLIENT] Ошибка парсинга или обработки: {e}", exc_info=True)
        raise HTTPException(500, "Ошибка обработки файла")


def generate_excel_from_models(data_list: List[BaseModel]) -> io.BytesIO:
    """
    Генерирует Excel файл из списка Pydantic моделей.
    Применяет форматирование дат ДД.ММ.ГГГГ.
    """
    output_stream = io.BytesIO()
    work_book = Workbook()
    sheet = work_book.active

    titles = ["ФИО", "ДР", "Возраст", "Адрес", "Страховая", "Номер полиса", "Номер карты",
              "Поступление", "Выписка", "Результат", "Койко-дни", "Отделение", "Профиль", "МКБ",
              "Диагноз", "Врач", "", "Код услуги", "Название", "Кол-во", "Дата", "Источник оплаты"]
    sheet.append(titles)


    # Если список пуст, возвращаем пустой файл (или можно добавить заголовки)
    if not data_list:
        work_book.save(output_stream)
        output_stream.seek(0)
        return output_stream

    # Логика заполнения
    for row in data_list:
        # 1. Превращаем модель в словарь и берем значения
        row_dict = row.model_dump()
        row_values = list(row_dict.values())

        # 2. Добавляем строку
        sheet.append(row_values)

        # 3. Форматирование дат
        current_row = sheet[sheet.max_row]
        for cell in current_row:
            if isinstance(cell.value, (date, datetime)):
                cell.number_format = 'DD.MM.YYYY'

    # автоширина всех колонок
    auto_cells_width(sheet)

    # задаем ширину колонок принудительно
    set_column_width(sheet=sheet, column_letter="D", width=45)
    set_column_width(sheet=sheet, column_letter="E", width=45)
    set_column_width(sheet=sheet, column_letter="L", width=45)
    set_column_width(sheet=sheet, column_letter="M", width=45)
    set_column_width(sheet=sheet, column_letter="Q", width=45)
    set_column_width(sheet=sheet, column_letter="S", width=45)

    # выравнивание первой строки по центру
    rows_to_align = [1]
    align_row_center(sheet, rows_to_align)

    # выравнивание колонок по центру
    columns_to_align = ["B", "C", "H", "I", "K", "T", "U"]
    align_column_center(sheet, columns_to_align)

    # включить АВТОФИЛЬТР
    # sheet.dimensions вернет строку вида "A1:K150"
    sheet.auto_filter.ref = sheet.dimensions

    # Сохраняем
    work_book.save(output_stream)
    output_stream.seek(0)

    return output_stream
